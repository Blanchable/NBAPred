"""
Excel Export for NBA Prediction Engine.

Exports predictions with performance summary header to Excel format.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.storage import (
    load_predictions_log,
    PredictionLogEntry,
    compute_performance_summary,
    OUTPUTS_DIR,
)


def export_predictions_to_excel(
    entries: List[PredictionLogEntry] = None,
    output_path: Path = None,
    title: str = "NBA Predictions",
) -> Path:
    """
    Export predictions to Excel with performance summary header.
    
    Args:
        entries: List of prediction entries (loads from log if None)
        output_path: Output file path (auto-generated if None)
        title: Title for the report
        
    Returns:
        Path to created Excel file
    """
    # Load entries if not provided
    if entries is None:
        entries = load_predictions_log()
    
    if not entries:
        raise ValueError("No predictions to export")
    
    # Create output path
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUTS_DIR / f"predictions_export_{timestamp}.xlsx"
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Compute performance summary
    summary = compute_performance_summary()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    stat_font = Font(size=11)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== HEADER SECTION ====================
    row = 1
    
    # Title
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    
    # Generated timestamp
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    row += 2
    
    # Performance Summary Header
    ws.cell(row=row, column=1, value="PERFORMANCE SUMMARY").font = header_font
    row += 1
    
    # Overall record
    ws.cell(row=row, column=1, value="Overall Record:").font = subheader_font
    overall_text = f"{summary.wins}/{summary.total_games}"
    ws.cell(row=row, column=2, value=overall_text)
    ws.cell(row=row, column=3, value=f"({summary.win_pct:.1%})" if summary.total_games > 0 else "")
    row += 2
    
    # Win % by Confidence - Header row
    ws.cell(row=row, column=1, value="Win % by Confidence Level:").font = subheader_font
    row += 1
    
    # Column headers for confidence table
    conf_headers = ["Confidence", "Record", "Win %"]
    for col, header in enumerate(conf_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # HIGH confidence
    high_record = f"{summary.high_conf_wins}/{summary.high_conf_games}"
    high_pct = f"{summary.high_conf_win_pct:.1%}" if summary.high_conf_games > 0 else "N/A"
    ws.cell(row=row, column=1, value="HIGH").border = thin_border
    ws.cell(row=row, column=1).fill = green_fill
    ws.cell(row=row, column=2, value=high_record).border = thin_border
    ws.cell(row=row, column=3, value=high_pct).border = thin_border
    row += 1
    
    # MEDIUM confidence
    med_record = f"{summary.med_conf_wins}/{summary.med_conf_games}"
    med_pct = f"{summary.med_conf_win_pct:.1%}" if summary.med_conf_games > 0 else "N/A"
    ws.cell(row=row, column=1, value="MEDIUM").border = thin_border
    ws.cell(row=row, column=1).fill = yellow_fill
    ws.cell(row=row, column=2, value=med_record).border = thin_border
    ws.cell(row=row, column=3, value=med_pct).border = thin_border
    row += 1
    
    # LOW confidence
    low_record = f"{summary.low_conf_wins}/{summary.low_conf_games}"
    low_pct = f"{summary.low_conf_win_pct:.1%}" if summary.low_conf_games > 0 else "N/A"
    ws.cell(row=row, column=1, value="LOW").border = thin_border
    ws.cell(row=row, column=1).fill = red_fill
    ws.cell(row=row, column=2, value=low_record).border = thin_border
    ws.cell(row=row, column=3, value=low_pct).border = thin_border
    row += 2
    
    # Recent performance
    ws.cell(row=row, column=1, value="Recent Performance:").font = subheader_font
    row += 1
    ws.cell(row=row, column=1, value="Last 7 Days:")
    ws.cell(row=row, column=2, value=f"{summary.last_7_days_wins}/{summary.last_7_days_games}")
    ws.cell(row=row, column=3, value=f"({summary.last_7_days_win_pct:.1%})" if summary.last_7_days_games > 0 else "")
    row += 1
    ws.cell(row=row, column=1, value="Last 30 Days:")
    ws.cell(row=row, column=2, value=f"{summary.last_30_days_wins}/{summary.last_30_days_games}")
    ws.cell(row=row, column=3, value=f"({summary.last_30_days_win_pct:.1%})" if summary.last_30_days_games > 0 else "")
    row += 3
    
    # ==================== PREDICTIONS TABLE ====================
    ws.cell(row=row, column=1, value="PREDICTIONS").font = header_font
    row += 1
    
    # Column headers
    columns = ["Date", "Matchup", "Pick", "Edge Score", "Confidence %", "Confidence Level", "Actual Winner", "Correct"]
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Sort entries by date descending, then by edge score
    sorted_entries = sorted(entries, key=lambda e: (e.game_date, -abs(e.edge_score_total)), reverse=True)
    
    # Data rows
    for entry in sorted_entries:
        matchup = f"{entry.away_team} @ {entry.home_team}"
        
        # Date
        ws.cell(row=row, column=1, value=entry.game_date).border = thin_border
        
        # Matchup
        ws.cell(row=row, column=2, value=matchup).border = thin_border
        
        # Pick
        ws.cell(row=row, column=3, value=entry.pick).border = thin_border
        
        # Edge Score
        cell = ws.cell(row=row, column=4, value=entry.edge_score_total)
        cell.border = thin_border
        cell.number_format = '+0.0;-0.0;0.0'
        
        # Confidence %
        ws.cell(row=row, column=5, value=entry.confidence_pct).border = thin_border
        
        # Confidence Level with color
        conf_cell = ws.cell(row=row, column=6, value=entry.confidence_level)
        conf_cell.border = thin_border
        conf_cell.alignment = Alignment(horizontal='center')
        if entry.confidence_level.upper() == "HIGH":
            conf_cell.fill = green_fill
        elif entry.confidence_level.upper() == "MEDIUM":
            conf_cell.fill = yellow_fill
        else:
            conf_cell.fill = red_fill
        
        # Actual Winner
        ws.cell(row=row, column=7, value=entry.actual_winner).border = thin_border
        
        # Correct
        correct_cell = ws.cell(row=row, column=8, value=entry.correct)
        correct_cell.border = thin_border
        correct_cell.alignment = Alignment(horizontal='center')
        if entry.correct == "1":
            correct_cell.fill = green_fill
            correct_cell.value = "✓"
        elif entry.correct == "0":
            correct_cell.fill = red_fill
            correct_cell.value = "✗"
        
        row += 1
    
    # Auto-adjust column widths
    column_widths = [12, 18, 8, 12, 14, 16, 14, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save workbook
    wb.save(output_path)
    
    return output_path


def export_season_backfill_to_excel(
    season: str = None,
    output_path: Path = None,
) -> Path:
    """
    Run full season backfill and export to Excel.
    
    Args:
        season: Season string (e.g., "2024-25"). Default: current season
        output_path: Output file path
        
    Returns:
        Path to created Excel file
    """
    from utils.dates import get_eastern_date, get_season_for_date
    
    # Determine season
    if season is None:
        season = get_season_for_date(get_eastern_date())
    
    # Load existing predictions
    entries = load_predictions_log()
    
    if not entries:
        raise ValueError("No predictions in log. Run backfill first.")
    
    # Filter to current season dates
    # Season 2024-25 runs from Oct 2024 to Jun 2025
    season_start_year = int(season.split("-")[0])
    season_start = f"{season_start_year}-10-01"
    season_end = f"{season_start_year + 1}-06-30"
    
    season_entries = [
        e for e in entries 
        if season_start <= e.game_date <= season_end
    ]
    
    if not season_entries:
        raise ValueError(f"No predictions found for season {season}")
    
    # Generate output path
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUTS_DIR / f"season_{season}_predictions_{timestamp}.xlsx"
    
    return export_predictions_to_excel(
        entries=season_entries,
        output_path=output_path,
        title=f"NBA Predictions - {season} Season",
    )
