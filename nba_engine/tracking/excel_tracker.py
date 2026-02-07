"""
Excel Tracking Module for NBA Prediction Engine.

Manages a single persistent Excel workbook with:
- LOG sheet: Main picks table with daily overwrite behavior
- STATS sheet: Auto-calculated winrate statistics by bucket
- SETTINGS sheet: Configuration thresholds and version info

The engine ONLY supports today's slate - no historical/backfill functionality.

IMPORTANT: The tracking file is stored in a PERSISTENT location that survives
app restarts, including when running as a PyInstaller frozen executable.

Location:
  Windows: %APPDATA%\\NBA_Engine\\tracking\\NBA_Engine_Tracking.xlsx
  macOS:   ~/Library/Application Support/NBA_Engine/tracking/
  Linux:   ~/.local/share/NBA_Engine/tracking/
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# Import persistent paths from the paths module
# This ensures the tracking file is written to a stable location
# that persists even when running as a frozen PyInstaller executable
from paths import (
    TRACKING_DIR,
    TRACKING_FILE_PATH,
    get_tracking_path_message,
)

# Sheet names
LOG_SHEET = "LOG"
STATS_SHEET = "STATS"
SETTINGS_SHEET = "SETTINGS"

# Column headers for LOG sheet (exact order per spec)
LOG_COLUMNS = [
    "Date",               # A - YYYY-MM-DD
    "Game_ID",            # B - string or blank
    "Away",               # C - away team abbrev
    "Home",               # D - home team abbrev
    "Pick",               # E - picked team
    "Side",               # F - HOME or AWAY
    "Conf_%",             # G - confidence percentage (float)
    "Bucket",             # H - HIGH/MED/LOW
    "Model_Prob",         # I - 0-1 probability
    "Edge",               # J - edge score
    "Pred_Away",          # K - predicted away points
    "Pred_Home",          # L - predicted home points
    "Pred_Total",         # M - predicted total
    "Total_Low",          # N - total range low
    "Total_High",         # O - total range high
    "Exp_Pace",           # P - expected possessions
    "PPP_Away",           # Q - away PPP
    "PPP_Home",           # R - home PPP
    "Var_Band",           # S - variance band width
    "Act_Away",           # T - actual away score (user fills)
    "Act_Home",           # U - actual home score (user fills)
    "Result",             # V - W/L (auto-calc or user fills)
    "Notes",              # W - optional
]

# Colors for bucket conditional formatting
COLORS = {
    'high_fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
    'high_font': Font(color="006100", bold=True),
    'med_fill': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # Light yellow/amber
    'med_font': Font(color="9C6500", bold=True),
    'low_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),   # Light red
    'low_font': Font(color="9C0006", bold=True),
    'win_fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'loss_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    'header_font': Font(color="FFFFFF", bold=True, size=11),
    'alt_row_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

# Styles
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


@dataclass
class PickEntry:
    """A single prediction entry for the LOG sheet."""
    run_date: str
    game_id: str
    away_team: str
    home_team: str
    pick_team: str
    pick_side: str           # "HOME" or "AWAY"
    confidence_pct: float    # e.g., 71.3
    confidence_bucket: str   # "HIGH", "MED", "LOW"
    model_prob: float        # 0-1 probability
    edge_score: float
    
    # Totals prediction fields
    pred_away_pts: int = 0      # Predicted away points
    pred_home_pts: int = 0      # Predicted home points
    pred_total: int = 0         # Predicted total
    total_range_low: int = 0    # Total range low
    total_range_high: int = 0   # Total range high
    expected_pace: float = 0.0  # Expected possessions
    ppp_away: float = 0.0       # Away PPP
    ppp_home: float = 0.0       # Home PPP
    variance_band: int = 12     # Variance band width
    
    # Actual scores (user fills)
    actual_away: str = ""
    actual_home: str = ""
    result: str = ""         # W/L - auto-calc or user fills
    notes: str = ""
    
    # Legacy compatibility fields
    run_timestamp: str = ""
    confidence_level: str = ""
    edge_score_total: float = 0.0
    projected_margin_home: float = 0.0
    home_win_prob: float = 0.0
    away_win_prob: float = 0.0
    top_5_factors: str = ""
    data_confidence: str = ""
    actual_winner: str = ""
    
    def to_row(self, row_num: int) -> list:
        """Convert to row values with formula for Result column."""
        # Auto-calc Result formula based on actual scores (columns T and U)
        # =IF(OR(T{row}="",U{row}=""),"",IF(AND(F{row}="HOME",U{row}>T{row}),"W",IF(AND(F{row}="AWAY",T{row}>U{row}),"W","L")))
        result_formula = f'=IF(OR(T{row_num}="",U{row_num}=""),"",IF(AND(F{row_num}="HOME",U{row_num}>T{row_num}),"W",IF(AND(F{row_num}="AWAY",T{row_num}>U{row_num}),"W","L")))'
        
        return [
            self.run_date,          # A
            self.game_id,           # B
            self.away_team,         # C
            self.home_team,         # D
            self.pick_team,         # E
            self.pick_side,         # F
            self.confidence_pct,    # G
            self.confidence_bucket, # H
            round(self.model_prob, 3),  # I
            round(self.edge_score, 2),  # J
            self.pred_away_pts,     # K
            self.pred_home_pts,     # L
            self.pred_total,        # M
            self.total_range_low,   # N
            self.total_range_high,  # O
            round(self.expected_pace, 1),  # P
            round(self.ppp_away, 3),       # Q
            round(self.ppp_home, 3),       # R
            self.variance_band,     # S
            self.actual_away,       # T
            self.actual_home,       # U
            result_formula,         # V - Formula for auto-calc
            self.notes,             # W
        ]


@dataclass
class WinrateStats:
    """Winrate statistics computed from the LOG sheet."""
    # Overall
    total_graded: int = 0
    wins: int = 0
    losses: int = 0
    win_pct: float = 0.0
    
    # By confidence bucket
    high_graded: int = 0
    high_wins: int = 0
    high_losses: int = 0
    high_win_pct: float = 0.0
    
    medium_graded: int = 0
    medium_wins: int = 0
    medium_losses: int = 0
    medium_win_pct: float = 0.0
    
    low_graded: int = 0
    low_wins: int = 0
    low_losses: int = 0
    low_win_pct: float = 0.0
    
    # Pending (no result yet)
    pending_total: int = 0
    pending_high: int = 0
    pending_medium: int = 0
    pending_low: int = 0
    
    # Counts
    total_high: int = 0
    total_medium: int = 0
    total_low: int = 0


class ExcelTracker:
    """
    Manages the NBA Engine tracking Excel workbook.
    
    Key features:
    - Single persistent workbook at outputs/tracking/NBA_Engine_Tracking.xlsx
    - Overwrite-by-day: Running multiple times replaces that day's picks
    - LOG sheet with proper table formatting and conditional formatting
    - STATS sheet with auto-calculated winrate stats
    - SETTINGS sheet with configuration thresholds
    """
    
    def __init__(self, file_path: Path = None):
        """Initialize the tracker."""
        self.file_path = file_path or TRACKING_FILE_PATH
        self._ensure_directory()
    
    def _ensure_directory(self):
        """Ensure the tracking directory exists."""
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
    
    def _create_new_workbook(self) -> Workbook:
        """Create a new workbook with LOG, STATS, and SETTINGS sheets."""
        wb = Workbook()
        
        # Rename default sheet to LOG
        log_sheet = wb.active
        log_sheet.title = LOG_SHEET
        self._initialize_log_sheet(log_sheet)
        
        # Create STATS sheet
        stats_sheet = wb.create_sheet(STATS_SHEET)
        self._initialize_stats_sheet(stats_sheet)
        
        # Create SETTINGS sheet
        settings_sheet = wb.create_sheet(SETTINGS_SHEET)
        self._initialize_settings_sheet(settings_sheet)
        
        return wb
    
    def _initialize_log_sheet(self, sheet):
        """Initialize the LOG sheet with headers and formatting."""
        # Add header row with styling
        for col_idx, header in enumerate(LOG_COLUMNS, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
        
        # Set column widths
        column_widths = {
            'A': 12,   # Date
            'B': 12,   # Game_ID
            'C': 6,    # Away
            'D': 6,    # Home
            'E': 6,    # Pick
            'F': 7,    # Side
            'G': 8,    # Conf_%
            'H': 8,    # Bucket
            'I': 10,   # Model_Prob
            'J': 7,    # Edge
            'K': 10,   # Pred_Away
            'L': 10,   # Pred_Home
            'M': 10,   # Pred_Total
            'N': 10,   # Total_Low
            'O': 10,   # Total_High
            'P': 10,   # Exp_Pace
            'Q': 10,   # PPP_Away
            'R': 10,   # PPP_Home
            'S': 10,   # Var_Band
            'T': 10,   # Act_Away
            'U': 10,   # Act_Home
            'V': 8,    # Result
            'W': 20,   # Notes
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Set row height for header
        sheet.row_dimensions[1].height = 25
        
        # Add data validation for Result column V (W or L only)
        result_dv = DataValidation(
            type="list",
            formula1='"W,L"',
            allow_blank=True
        )
        result_dv.error = "Please enter W or L"
        result_dv.errorTitle = "Invalid Result"
        sheet.add_data_validation(result_dv)
        result_dv.add('V2:V1000')
        
        # Enable auto-filter
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(LOG_COLUMNS))}1"
    
    def _initialize_stats_sheet(self, sheet):
        """Initialize the STATS sheet with formulas pulling from LOG."""
        # Title styling
        title_font = Font(bold=True, size=16, color="4472C4")
        section_font = Font(bold=True, size=12, color="4472C4")
        label_font = Font(size=11)
        value_font = Font(size=11, bold=True)
        
        # Title
        sheet['A1'] = "NBA Engine Performance Dashboard"
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:D1')
        
        sheet['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A2'].font = Font(size=10, italic=True, color="666666")
        
        # Overall Performance Section
        sheet['A4'] = "OVERALL PERFORMANCE"
        sheet['A4'].font = section_font
        
        labels = [
            ('A5', 'Total Picks:'),
            ('A6', 'Graded:'),
            ('A7', 'Wins:'),
            ('A8', 'Losses:'),
            ('A9', 'Win Rate:'),
            ('A10', 'Pending:'),
        ]
        for cell, label in labels:
            sheet[cell] = label
            sheet[cell].font = label_font
        
        # Formulas for overall stats (reference LOG sheet, Result is column V)
        sheet['B5'] = '=COUNTA(LOG!A:A)-1'  # Total picks
        sheet['B6'] = '=COUNTIF(LOG!V:V,"W")+COUNTIF(LOG!V:V,"L")'  # Graded
        sheet['B7'] = '=COUNTIF(LOG!V:V,"W")'  # Wins
        sheet['B8'] = '=COUNTIF(LOG!V:V,"L")'  # Losses
        sheet['B9'] = '=IF(B6>0,B7/B6,0)'  # Win Rate
        sheet['B9'].number_format = '0.0%'
        sheet['B10'] = '=B5-B6'  # Pending
        
        for row in range(5, 11):
            sheet[f'B{row}'].font = value_font
        
        # By Bucket Section
        sheet['A12'] = "BY CONFIDENCE BUCKET"
        sheet['A12'].font = section_font
        
        # Headers
        headers = ['Bucket', 'Total', 'Graded', 'Wins', 'Losses', 'Win %', 'Pending']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=13, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # HIGH row (Result is now column V)
        sheet['A14'] = 'HIGH'
        sheet['A14'].font = COLORS['high_font']
        sheet['A14'].fill = COLORS['high_fill']
        sheet['B14'] = '=COUNTIF(LOG!H:H,"HIGH")'
        sheet['C14'] = '=SUMPRODUCT((LOG!H:H="HIGH")*(LOG!V:V<>""))'
        sheet['D14'] = '=SUMPRODUCT((LOG!H:H="HIGH")*(LOG!V:V="W"))'
        sheet['E14'] = '=SUMPRODUCT((LOG!H:H="HIGH")*(LOG!V:V="L"))'
        sheet['F14'] = '=IF(C14>0,D14/C14,0)'
        sheet['F14'].number_format = '0.0%'
        sheet['G14'] = '=B14-C14'
        
        # MED row (Result is now column V)
        sheet['A15'] = 'MEDIUM'
        sheet['A15'].font = COLORS['med_font']
        sheet['A15'].fill = COLORS['med_fill']
        sheet['B15'] = '=COUNTIF(LOG!H:H,"MED")+COUNTIF(LOG!H:H,"MEDIUM")'
        sheet['C15'] = '=SUMPRODUCT((LOG!H:H="MED")*(LOG!V:V<>""))+SUMPRODUCT((LOG!H:H="MEDIUM")*(LOG!V:V<>""))'
        sheet['D15'] = '=SUMPRODUCT((LOG!H:H="MED")*(LOG!V:V="W"))+SUMPRODUCT((LOG!H:H="MEDIUM")*(LOG!V:V="W"))'
        sheet['E15'] = '=SUMPRODUCT((LOG!H:H="MED")*(LOG!V:V="L"))+SUMPRODUCT((LOG!H:H="MEDIUM")*(LOG!V:V="L"))'
        sheet['F15'] = '=IF(C15>0,D15/C15,0)'
        sheet['F15'].number_format = '0.0%'
        sheet['G15'] = '=B15-C15'
        
        # LOW row (Result is now column V)
        sheet['A16'] = 'LOW'
        sheet['A16'].font = COLORS['low_font']
        sheet['A16'].fill = COLORS['low_fill']
        sheet['B16'] = '=COUNTIF(LOG!H:H,"LOW")'
        sheet['C16'] = '=SUMPRODUCT((LOG!H:H="LOW")*(LOG!V:V<>""))'
        sheet['D16'] = '=SUMPRODUCT((LOG!H:H="LOW")*(LOG!V:V="W"))'
        sheet['E16'] = '=SUMPRODUCT((LOG!H:H="LOW")*(LOG!V:V="L"))'
        sheet['F16'] = '=IF(C16>0,D16/C16,0)'
        sheet['F16'].number_format = '0.0%'
        sheet['G16'] = '=B16-C16'
        
        # Center align all data cells
        for row in range(14, 17):
            for col in range(1, 8):
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        # Column widths
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        
        # Recent Performance Section
        sheet['A18'] = "QUICK STATS"
        sheet['A18'].font = section_font
        
        sheet['A19'] = "Today's Picks:"
        sheet['B19'] = f'=COUNTIF(LOG!A:A,"{datetime.now().strftime("%Y-%m-%d")}")'
        sheet['A20'] = "This Week:"
        sheet['B20'] = '=COUNTIF(LOG!A:A,">="&TODAY()-7)'
    
    def _initialize_settings_sheet(self, sheet):
        """Initialize the SETTINGS sheet with configuration."""
        sheet['A1'] = "NBA Engine Configuration"
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = "Confidence Thresholds"
        sheet['A3'].font = Font(bold=True)
        
        sheet['A4'] = "HIGH minimum:"
        sheet['B4'] = 65.0
        sheet['C4'] = "%(confidence >= this is HIGH)"
        
        sheet['A5'] = "MEDIUM minimum:"
        sheet['B5'] = 57.5
        sheet['C5'] = "%(confidence >= this is MEDIUM, else LOW)"
        
        sheet['A7'] = "Version Info"
        sheet['A7'].font = Font(bold=True)
        sheet['A8'] = "Engine Version:"
        sheet['B8'] = "v3.1"
        sheet['A9'] = "Tracker Version:"
        sheet['B9'] = "2.0"
        sheet['A10'] = "Created:"
        sheet['B10'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 40
    
    def _apply_log_formatting(self, sheet, data_start: int, data_end: int):
        """Apply conditional formatting and styling to LOG data rows."""
        if data_end < data_start:
            return
        
        # Bucket column conditional formatting (column H)
        bucket_col = 'H'
        
        # HIGH - green
        sheet.conditional_formatting.add(
            f'{bucket_col}{data_start}:{bucket_col}{data_end}',
            FormulaRule(
                formula=[f'{bucket_col}{data_start}="HIGH"'],
                fill=COLORS['high_fill'],
                font=COLORS['high_font']
            )
        )
        
        # MED/MEDIUM - yellow
        sheet.conditional_formatting.add(
            f'{bucket_col}{data_start}:{bucket_col}{data_end}',
            FormulaRule(
                formula=[f'OR({bucket_col}{data_start}="MED",{bucket_col}{data_start}="MEDIUM")'],
                fill=COLORS['med_fill'],
                font=COLORS['med_font']
            )
        )
        
        # LOW - red
        sheet.conditional_formatting.add(
            f'{bucket_col}{data_start}:{bucket_col}{data_end}',
            FormulaRule(
                formula=[f'{bucket_col}{data_start}="LOW"'],
                fill=COLORS['low_fill'],
                font=COLORS['low_font']
            )
        )
        
        # Result column conditional formatting (column V)
        result_col = 'V'
        
        # W - green
        sheet.conditional_formatting.add(
            f'{result_col}{data_start}:{result_col}{data_end}',
            FormulaRule(
                formula=[f'{result_col}{data_start}="W"'],
                fill=COLORS['win_fill'],
                font=Font(color="006100", bold=True)
            )
        )
        
        # L - red
        sheet.conditional_formatting.add(
            f'{result_col}{data_start}:{result_col}{data_end}',
            FormulaRule(
                formula=[f'{result_col}{data_start}="L"'],
                fill=COLORS['loss_fill'],
                font=Font(color="9C0006", bold=True)
            )
        )
        
        # Bold the predicted total column (column M)
        for row_idx in range(data_start, data_end + 1):
            sheet.cell(row=row_idx, column=13).font = Font(bold=True)  # Pred_Total
        
        # Apply row styling
        for row_idx in range(data_start, data_end + 1):
            # Alternate row background
            if row_idx % 2 == 0:
                for col_idx in range(1, len(LOG_COLUMNS) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.fill.start_color.rgb == '00000000' or cell.fill.start_color.rgb is None:
                        cell.fill = COLORS['alt_row_fill']
            
            # Center align team columns
            for col in ['C', 'D', 'E', 'F', 'H', 'M']:
                col_idx = ord(col) - ord('A') + 1
                sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
            
            # Right align numeric columns
            for col in ['G', 'I', 'J', 'K', 'L']:
                col_idx = ord(col) - ord('A') + 1
                sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='right')
            
            # Format confidence % with 1 decimal
            sheet.cell(row=row_idx, column=7).number_format = '0.0'
            
            # Format model prob with 3 decimals
            sheet.cell(row=row_idx, column=9).number_format = '0.000'
    
    def _load_or_create_workbook(self) -> Workbook:
        """Load existing workbook or create new one."""
        if self.file_path.exists():
            try:
                return load_workbook(self.file_path)
            except Exception as e:
                raise IOError(
                    f"Cannot open {self.file_path.name}. "
                    f"Please close the file and try again.\nError: {e}"
                )
        else:
            return self._create_new_workbook()
    
    def save_predictions(self, picks: List[PickEntry]) -> int:
        """
        Save predictions to the LOG sheet.
        
        Implements OVERWRITE-BY-DAY rule:
        - Deletes all rows where Date == today's date
        - Appends new predictions at the bottom
        
        Args:
            picks: List of PickEntry objects to save
            
        Returns:
            Number of picks saved
        """
        if not picks:
            return 0
        
        # Get today's date
        today = picks[0].run_date
        
        try:
            wb = self._load_or_create_workbook()
        except IOError as e:
            raise e
        
        # Ensure LOG sheet exists
        if LOG_SHEET not in wb.sheetnames:
            log_sheet = wb.create_sheet(LOG_SHEET, 0)
            self._initialize_log_sheet(log_sheet)
        else:
            log_sheet = wb[LOG_SHEET]
        
        # Find and delete all rows for today's date
        rows_to_delete = []
        for row_idx in range(2, log_sheet.max_row + 1):
            cell_value = log_sheet.cell(row=row_idx, column=1).value
            if cell_value == today:
                rows_to_delete.append(row_idx)
        
        # Delete rows in reverse order to maintain indices
        for row_idx in reversed(rows_to_delete):
            log_sheet.delete_rows(row_idx)
        
        # Find the next available row
        next_row = log_sheet.max_row + 1
        if next_row == 1:  # Empty sheet (only header exists but max_row is 1)
            next_row = 2
        
        data_start = next_row
        
        # Append new predictions
        for pick in picks:
            row_data = pick.to_row(next_row)
            for col_idx, value in enumerate(row_data, 1):
                cell = log_sheet.cell(row=next_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
            next_row += 1
        
        data_end = next_row - 1
        
        # Apply formatting to new rows
        self._apply_log_formatting(log_sheet, data_start, data_end)
        
        # Update auto-filter range
        log_sheet.auto_filter.ref = f"A1:{get_column_letter(len(LOG_COLUMNS))}{log_sheet.max_row}"
        
        # Update STATS sheet timestamp
        if STATS_SHEET in wb.sheetnames:
            wb[STATS_SHEET]['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Save workbook
        try:
            wb.save(self.file_path)
        except PermissionError:
            raise IOError(
                f"Please close {self.file_path.name} and try again."
            )
        
        return len(picks)
    
    def compute_winrate_stats(self) -> WinrateStats:
        """
        Compute winrate statistics from the LOG sheet.
        
        Computes W/L in Python from actual scores (Act_Away, Act_Home) and Side,
        rather than relying on Excel formula caching. This ensures the GUI shows
        correct win rates immediately without requiring Excel to be opened first.
        
        Returns:
            WinrateStats object with computed statistics
        """
        stats = WinrateStats()
        
        if not self.file_path.exists():
            return stats
        
        try:
            # Use data_only=False to read the raw values, not cached formula results
            wb = load_workbook(self.file_path, data_only=False)
        except Exception:
            return stats
        
        if LOG_SHEET not in wb.sheetnames:
            return stats
        
        log_sheet = wb[LOG_SHEET]
        
        def _parse_int(val) -> Optional[int]:
            """Safely parse a value to int, return None if not parseable."""
            if val is None:
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None
        
        def _compute_result(side: str, act_away: Optional[int], act_home: Optional[int], result_cell) -> Optional[str]:
            """
            Compute game result from actual scores and side.
            
            Priority:
            1. If result_cell is explicitly 'W' or 'L', use it
            2. If actual scores are present, compute from Side and scores
            3. Otherwise return None (pending)
            """
            # Check if result_cell has an explicit W/L (not a formula)
            if result_cell and isinstance(result_cell, str):
                result_upper = result_cell.strip().upper()
                if result_upper in ('W', 'L'):
                    return result_upper
            
            # Compute from actual scores if available
            if act_away is not None and act_home is not None and side:
                side_upper = str(side).strip().upper()
                if side_upper == "HOME":
                    return "W" if act_home > act_away else "L"
                elif side_upper == "AWAY":
                    return "W" if act_away > act_home else "L"
            
            return None  # Pending
        
        # Process each data row
        for row_idx in range(2, log_sheet.max_row + 1):
            # Column indices (1-indexed):
            # E=5 (Pick), F=6 (Side), H=8 (Bucket)
            # T=20 (Act_Away), U=21 (Act_Home), V=22 (Result)
            bucket = log_sheet.cell(row=row_idx, column=8).value  # H - Bucket
            pick_team = log_sheet.cell(row=row_idx, column=5).value  # E - Pick
            side = log_sheet.cell(row=row_idx, column=6).value  # F - Side
            act_away = _parse_int(log_sheet.cell(row=row_idx, column=20).value)  # T - Act_Away
            act_home = _parse_int(log_sheet.cell(row=row_idx, column=21).value)  # U - Act_Home
            result_cell = log_sheet.cell(row=row_idx, column=22).value  # V - Result
            
            if not bucket or not pick_team:
                continue
            
            bucket = str(bucket).upper()
            
            # Normalize bucket name
            if bucket == "MEDIUM":
                bucket = "MED"
            
            # Count by bucket
            if bucket == "HIGH":
                stats.total_high += 1
            elif bucket == "MED":
                stats.total_medium += 1
            elif bucket == "LOW":
                stats.total_low += 1
            
            # Compute result (W/L or None for pending)
            result_final = _compute_result(side, act_away, act_home, result_cell)
            
            # Check if graded or pending
            if result_final in ('W', 'L'):
                is_win = (result_final == 'W')
                
                stats.total_graded += 1
                if is_win:
                    stats.wins += 1
                else:
                    stats.losses += 1
                
                # By bucket
                if bucket == "HIGH":
                    stats.high_graded += 1
                    if is_win:
                        stats.high_wins += 1
                    else:
                        stats.high_losses += 1
                elif bucket == "MED":
                    stats.medium_graded += 1
                    if is_win:
                        stats.medium_wins += 1
                    else:
                        stats.medium_losses += 1
                elif bucket == "LOW":
                    stats.low_graded += 1
                    if is_win:
                        stats.low_wins += 1
                    else:
                        stats.low_losses += 1
            else:
                # Pending pick
                stats.pending_total += 1
                if bucket == "HIGH":
                    stats.pending_high += 1
                elif bucket == "MED":
                    stats.pending_medium += 1
                elif bucket == "LOW":
                    stats.pending_low += 1
        
        # Calculate win percentages
        if stats.total_graded > 0:
            stats.win_pct = (stats.wins / stats.total_graded) * 100
        if stats.high_graded > 0:
            stats.high_win_pct = (stats.high_wins / stats.high_graded) * 100
        if stats.medium_graded > 0:
            stats.medium_win_pct = (stats.medium_wins / stats.medium_graded) * 100
        if stats.low_graded > 0:
            stats.low_win_pct = (stats.low_wins / stats.low_graded) * 100
        
        wb.close()
        return stats
    
    def update_summary_sheet(self, stats: WinrateStats = None):
        """
        Update the STATS sheet timestamp.
        
        Note: STATS sheet uses formulas to pull data from LOG,
        so we just need to update the timestamp.
        """
        if not self.file_path.exists():
            return
        
        try:
            wb = load_workbook(self.file_path)
        except Exception:
            return
        
        if STATS_SHEET in wb.sheetnames:
            wb[STATS_SHEET]['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        try:
            wb.save(self.file_path)
        except PermissionError:
            pass  # Silently fail if file is locked
        
        wb.close()
    
    def refresh_winrates(self) -> WinrateStats:
        """
        Refresh winrate statistics by re-reading the Excel file.
        
        Returns:
            Updated WinrateStats
        """
        stats = self.compute_winrate_stats()
        self.update_summary_sheet(stats)
        return stats
    
    def file_exists(self) -> bool:
        """Check if the tracking file exists."""
        return self.file_path.exists()
    
    def get_file_path(self) -> str:
        """Get the tracking file path as string."""
        return str(self.file_path)
